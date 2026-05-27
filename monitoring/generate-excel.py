#!/usr/bin/env python3
"""
generate-excel.py
=================
Reads the docker stats CSV produced by start-monitoring.sh and writes a
formatted Excel workbook with per-container averages and raw measurements.

Usage:
    python3 generate-excel.py [--data-dir DIR] [--output FILE]

Defaults:
    --data-dir   monitoring/data
    --output     monitoring/data/resource-report.xlsx

Sheets produced:
    Summary          one row per container, avg/max/min CPU & memory
    SUT Containers   same metrics, filtered to non-browser containers
    Browsers         chrome-node, chrome-video, selenium-hub
    Raw Data         every measurement row (sortable / filterable in Excel)
"""

import argparse
import csv
import os
import re
import sys
from collections import defaultdict
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import (
        Alignment, Border, Font, PatternFill, Side,
    )
    from openpyxl.utils import get_column_letter
except ImportError:
    print(
        "[ERROR] openpyxl is not installed.\n"
        "        Run:  pip install openpyxl\n"
        "        or:   pip install -r monitoring/requirements.txt",
        file=sys.stderr,
    )
    sys.exit(1)


# ── Unit conversion ──────────────────────────────────────────────────────────

_UNIT_MIB = {
    "b":   1 / (1024 ** 2),
    "kb":  1 / 1024,       "kib": 1 / 1024,
    "mb":  1 / 1.048576,   "mib": 1.0,
    "gb":  1024 / 1.048576,"gib": 1024.0,
    "tb":  1024**2 / 1.048576, "tib": 1024**2,
}


def _to_mib(text: str) -> float:
    """'256MiB' → 256.0  |  '1.5GB' → 1430.51  |  '0B' → 0.0"""
    text = text.strip()
    m = re.match(r"^([\d.]+)\s*([A-Za-z]+)$", text)
    if not m:
        return 0.0
    value, unit = float(m.group(1)), m.group(2).lower()
    return value * _UNIT_MIB.get(unit, 0.0)


def _parse_mem(field: str):
    """'256MiB / 16GiB' → (usage_mib, limit_mib)"""
    parts = field.split("/")
    if len(parts) != 2:
        return 0.0, 0.0
    return _to_mib(parts[0]), _to_mib(parts[1])


def _parse_pct(text: str) -> float:
    """'2.50%' → 2.5  |  '--' → 0.0"""
    cleaned = text.strip().rstrip("%")
    try:
        return float(cleaned)
    except ValueError:
        return 0.0


# ── CSV loading ───────────────────────────────────────────────────────────────

_BROWSER_PATTERNS = re.compile(
    r"chrome.node|chrome.video|selenoid|selenium.hub",
    re.IGNORECASE,
)


def _is_browser(name: str) -> bool:
    return bool(_BROWSER_PATTERNS.search(name))


def load_data(data_dir: str):
    """
    Returns a list of dicts, one per measurement row:
        timestamp, container, cpu_pct, mem_usage_mib, mem_limit_mib,
        mem_pct, net_io, block_io, pids, is_browser
    """
    csv_path = Path(data_dir) / "stats.csv"
    if not csv_path.exists():
        return []

    rows = []
    with open(csv_path, newline="", encoding="utf-8") as fh:
        reader = csv.DictReader(fh)
        for raw in reader:
            try:
                usage_mib, limit_mib = _parse_mem(raw.get("mem_usage", "0B / 0B"))
                rows.append({
                    "timestamp":     raw.get("timestamp", ""),
                    "container":     raw.get("container", ""),
                    "cpu_pct":       _parse_pct(raw.get("cpu_pct", "0%")),
                    "mem_usage_mib": round(usage_mib, 2),
                    "mem_limit_mib": round(limit_mib, 2),
                    "mem_pct":       _parse_pct(raw.get("mem_pct", "0%")),
                    "net_io":        raw.get("net_io", ""),
                    "block_io":      raw.get("block_io", ""),
                    "pids":          raw.get("pids", ""),
                    "is_browser":    _is_browser(raw.get("container", "")),
                })
            except (ValueError, KeyError):
                pass  # skip malformed rows
    return rows


# ── Aggregation ───────────────────────────────────────────────────────────────

def _stats(values):
    if not values:
        return dict(avg=0.0, max=0.0, min=0.0, p95=0.0, count=0)
    s = sorted(values)
    n = len(s)
    p95_idx = max(0, int(n * 0.95) - 1)
    return dict(
        avg=round(sum(s) / n, 2),
        max=round(max(s), 2),
        min=round(min(s), 2),
        p95=round(s[p95_idx], 2),
        count=n,
    )


def aggregate(rows):
    """
    Returns list of dicts keyed by container name, each with cpu and mem stats.
    """
    groups = defaultdict(lambda: {"cpu": [], "mem": [], "mem_pct": [], "is_browser": False})
    for r in rows:
        g = groups[r["container"]]
        g["cpu"].append(r["cpu_pct"])
        g["mem"].append(r["mem_usage_mib"])
        g["mem_pct"].append(r["mem_pct"])
        g["is_browser"] = r["is_browser"]

    result = []
    for container, g in sorted(groups.items()):
        cpu = _stats(g["cpu"])
        mem = _stats(g["mem"])
        mem_pct = _stats(g["mem_pct"])
        result.append({
            "container":    container,
            "is_browser":   g["is_browser"],
            "cpu_avg":      cpu["avg"],
            "cpu_max":      cpu["max"],
            "cpu_min":      cpu["min"],
            "cpu_p95":      cpu["p95"],
            "mem_avg_mib":  mem["avg"],
            "mem_max_mib":  mem["max"],
            "mem_min_mib":  mem["min"],
            "mem_p95_mib":  mem["p95"],
            "mem_pct_avg":  mem_pct["avg"],
            "samples":      cpu["count"],
        })
    return result


# ── Excel styling ─────────────────────────────────────────────────────────────

_BLUE_DARK  = PatternFill(fill_type="solid", fgColor="1F497D")
_BLUE_LIGHT = PatternFill(fill_type="solid", fgColor="DCE6F1")
_GREEN_DARK = PatternFill(fill_type="solid", fgColor="375623")
_GREEN_ALT  = PatternFill(fill_type="solid", fgColor="E2EFDA")
_THIN       = Side(style="thin")
_BORDER     = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_HDR_FONT   = Font(bold=True, color="FFFFFF", size=11)
_BOLD       = Font(bold=True)


def _header_row(ws, headers, row=1, fill=_BLUE_DARK):
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = _HDR_FONT
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _BORDER
    ws.row_dimensions[row].height = 30


def _data_cell(ws, row, col, value, alt=False, fill_override=None):
    cell = ws.cell(row=row, column=col, value=value)
    if fill_override:
        cell.fill = fill_override
    elif alt:
        cell.fill = _BLUE_LIGHT
    cell.border = _BORDER
    cell.alignment = Alignment(horizontal="center")
    return cell


def _auto_width(ws, min_width=12, max_width=40):
    for col_cells in ws.columns:
        width = min_width
        for cell in col_cells:
            if cell.value:
                width = max(width, min(len(str(cell.value)) + 4, max_width))
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = width


def _freeze(ws, cell="A2"):
    ws.freeze_panes = cell


# ── Sheet writers ─────────────────────────────────────────────────────────────

_AGG_HEADERS = [
    "Container",
    "Avg CPU (%)", "Max CPU (%)", "Min CPU (%)", "P95 CPU (%)",
    "Avg Mem (MiB)", "Max Mem (MiB)", "Min Mem (MiB)", "P95 Mem (MiB)",
    "Avg Mem (%)",
    "Samples",
]


def _write_agg_sheet(ws, agg_rows, hdr_fill=_BLUE_DARK, alt_fill=_BLUE_LIGHT):
    _header_row(ws, _AGG_HEADERS, fill=hdr_fill)
    _freeze(ws)
    for i, r in enumerate(agg_rows, 2):
        alt = (i % 2 == 0)
        values = [
            r["container"],
            r["cpu_avg"], r["cpu_max"], r["cpu_min"], r["cpu_p95"],
            r["mem_avg_mib"], r["mem_max_mib"], r["mem_min_mib"], r["mem_p95_mib"],
            r["mem_pct_avg"],
            r["samples"],
        ]
        for col, val in enumerate(values, 1):
            f = alt_fill if alt else None
            _data_cell(ws, i, col, val, fill_override=f)
    _auto_width(ws)


def write_summary_sheet(wb, agg_rows):
    ws = wb.create_sheet("Summary")
    _write_agg_sheet(ws, agg_rows)


def write_sut_sheet(wb, agg_rows):
    ws = wb.create_sheet("SUT Containers")
    sut = [r for r in agg_rows if not r["is_browser"]]
    _write_agg_sheet(ws, sut)


def write_browser_sheet(wb, agg_rows):
    ws = wb.create_sheet("Browsers")
    browsers = [r for r in agg_rows if r["is_browser"]]
    if not browsers:
        ws.cell(row=1, column=1,
                value="No browser containers detected (chrome-node / chrome-video / selenium-hub).")
        ws.column_dimensions["A"].width = 70
        return
    _write_agg_sheet(ws, browsers, hdr_fill=_GREEN_DARK, alt_fill=_GREEN_ALT)


def write_raw_sheet(wb, rows):
    ws = wb.create_sheet("Raw Data")
    if not rows:
        ws.cell(row=1, column=1, value="No data collected.")
        return

    headers = [
        "Timestamp", "Container", "Type",
        "CPU (%)", "Mem Usage (MiB)", "Mem Limit (MiB)", "Mem (%)",
        "Net I/O", "Block I/O", "PIDs",
    ]
    _header_row(ws, headers)
    _freeze(ws)
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    for i, r in enumerate(rows, 2):
        alt = (i % 2 == 0)
        fill = _BLUE_LIGHT if alt else None
        values = [
            r["timestamp"],
            r["container"],
            "Browser" if r["is_browser"] else "SUT",
            r["cpu_pct"],
            r["mem_usage_mib"],
            r["mem_limit_mib"],
            r["mem_pct"],
            r["net_io"],
            r["block_io"],
            r["pids"],
        ]
        for col, val in enumerate(values, 1):
            _data_cell(ws, i, col, val, fill_override=fill)
    _auto_width(ws)


# ── Workbook assembly ─────────────────────────────────────────────────────────

def build_workbook(rows):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove the default empty sheet

    agg = aggregate(rows)
    write_summary_sheet(wb, agg)
    write_sut_sheet(wb, agg)
    write_browser_sheet(wb, agg)
    write_raw_sheet(wb, rows)
    return wb


# ── Entry point ───────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Generate an Excel resource-consumption report from docker stats CSV."
    )
    parser.add_argument(
        "--data-dir",
        default=str(Path(__file__).parent / "data"),
        help="Directory containing stats.csv (default: monitoring/data)",
    )
    parser.add_argument(
        "--output",
        default=None,
        help="Output .xlsx path (default: <data-dir>/resource-report.xlsx)",
    )
    args = parser.parse_args()

    output = args.output or str(Path(args.data_dir) / "resource-report.xlsx")

    print(f"[INFO] Reading data from {args.data_dir}/stats.csv ...")
    rows = load_data(args.data_dir)

    if not rows:
        print(
            "[WARN] No data found. Make sure start-monitoring.sh ran before the test suite.",
            file=sys.stderr,
        )
        sys.exit(1)

    containers = {r["container"] for r in rows}
    browsers   = {r["container"] for r in rows if r["is_browser"]}
    print(f"[INFO] {len(rows)} measurements across {len(containers)} containers "
          f"({len(browsers)} browser, {len(containers) - len(browsers)} SUT)")

    wb = build_workbook(rows)

    os.makedirs(os.path.dirname(os.path.abspath(output)), exist_ok=True)
    wb.save(output)
    print(f"[INFO] Excel report written → {output}")


if __name__ == "__main__":
    main()
